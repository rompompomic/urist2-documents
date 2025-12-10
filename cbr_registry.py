"""
Работа с XLSX-справочниками ЦБ РФ для актуализации данных о банках и МФО/МКК

Источники данных:
1. Банки: https://cbr.ru/banking_sector/credit/FullCoList/
   Кнопка "Экспортировать в XLSX"
   
2. МФО: https://cbr.ru/vfs/finmarkets/files/supervision/list_MFO.xlsx
   Прямая ссылка на XLSX

Преимущества:
- Актуальные данные напрямую от ЦБ РФ
- Не требует API ключей
- Простой формат XLSX
- Можно автоматически обновлять периодически
"""

import requests
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime


class CBRExcelRegistry:
    """Класс для работы со XLSX-справочниками ЦБ РФ"""
    
    # URL для скачивания справочников
    # Полный справочник с ИНН (более детальный)
    BANKS_FULL_URL = "https://cbr.ru/Queries/UniDbQuery/DownloadExcel/101215"
    # Базовый справочник (без ИНН)
    BANKS_URL = "https://cbr.ru/Queries/UniDbQuery/DownloadExcel/98547"
    MFO_URL = "https://cbr.ru/vfs/finmarkets/files/supervision/list_MFO.xlsx"
    
    # Папка для хранения скачанных файлов
    DATA_DIR = Path("cbr_data")
    
    def __init__(self):
        """Создает папку для данных, если её нет"""
        self.DATA_DIR.mkdir(exist_ok=True)
    
    def download_banks_registry(self) -> Optional[Path]:
        """
        Скачивает актуальный справочник банков (FullCoList)
        
        Returns:
            Path к скачанному файлу или None при ошибке
        """
        print("📥 Скачиваю справочник банков...")
        
        # Формируем URL с текущей датой
        today = datetime.now().strftime("%m/%d/%Y")
        url = f"{self.BANKS_URL}?FromDate={today}&ToDate={today}&posted=False"
        
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            # Сохраняем файл
            filepath = self.DATA_DIR / "banks_registry.xlsx"
            filepath.write_bytes(response.content)
            
            print(f"✅ Справочник банков сохранен: {filepath}")
            print(f"   Размер: {len(response.content) / 1024:.1f} KB")
            return filepath
            
        except requests.RequestException as e:
            print(f"❌ Ошибка скачивания справочника банков: {e}")
            return None
    
    def download_mfo_registry(self) -> Optional[Path]:
        """
        Скачивает актуальный справочник МФО
        
        Returns:
            Path к скачанному файлу или None при ошибке
        """
        print("\n📥 Скачиваю справочник МФО...")
        
        try:
            response = requests.get(self.MFO_URL, timeout=30)
            response.raise_for_status()
            
            # Сохраняем файл
            filepath = self.DATA_DIR / "mfo_registry.xlsx"
            filepath.write_bytes(response.content)
            
            print(f"✅ Справочник МФО сохранен: {filepath}")
            print(f"   Размер: {len(response.content) / 1024:.1f} KB")
            return filepath
            
        except requests.RequestException as e:
            print(f"❌ Ошибка скачивания справочника МФО: {e}")
            return None
    
    def parse_banks_registry(self, filepath: Path) -> List[Dict]:
        """
        Парсит XLSX-файл со справочником банков
        
        Args:
            filepath: Путь к XLSX-файлу
        
        Returns:
            Список словарей с данными банков
        """
        print(f"\n📖 Парсю справочник банков: {filepath}")
        
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            
            # Ищем заголовки (обычно в первой строке)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            print(f"   Найдено колонок: {len(headers)}")
            print(f"   Заголовки: {', '.join(str(h) for h in headers[:5])}...")
            
            # Читаем данные
            banks = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Пропускаем пустые строки
                    continue
                
                # Создаем словарь по заголовкам
                bank_data = {}
                for header, value in zip(headers, row):
                    if header and value:
                        bank_data[str(header).strip()] = str(value).strip()
                
                banks.append(bank_data)
                
                # Показываем прогресс
                if row_idx % 100 == 0:
                    print(f"   Обработано строк: {row_idx}")
            
            workbook.close()
            
            print(f"✅ Загружено банков: {len(banks)}")
            return banks
            
        except Exception as e:
            print(f"❌ Ошибка парсинга справочника банков: {e}")
            return []
    
    def parse_mfo_registry(self, filepath: Path) -> List[Dict]:
        """
        Парсит XLSX-файл со справочником МФО
        
        Args:
            filepath: Путь к XLSX-файлу
        
        Returns:
            Список словарей с данными МФО
        """
        print(f"\n📖 Парсю справочник МФО: {filepath}")
        
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            
            # В файле МФО заголовки находятся в строке 5
            headers = []
            for cell in sheet[5]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)
            
            print(f"   Найдено колонок: {len([h for h in headers if h])}")
            print(f"   Заголовки: {', '.join(h for h in headers[:5] if h)}...")
            
            # Читаем данные начиная с 6-й строки
            mfo_list = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), start=6):
                if not row or not row[0]:  # Пропускаем пустые строки
                    continue
                
                mfo_data = {}
                for header, value in zip(headers, row):
                    if header and value:
                        mfo_data[header] = str(value).strip()
                
                # Добавляем только если есть хотя бы название и ИНН
                if mfo_data.get('Полное наименование') or mfo_data.get('Сокращенное наименование'):
                    mfo_list.append(mfo_data)
                
                if row_idx % 100 == 0:
                    print(f"   Обработано строк: {row_idx}")
            
            workbook.close()
            
            print(f"✅ Загружено МФО: {len(mfo_list)}")
            
            # Показываем пример записи с ключевыми полями
            if mfo_list:
                example = mfo_list[0]
                print(f"\n   Пример МФО:")
                print(f"   Название: {example.get('Сокращенное наименование', 'N/A')}")
                print(f"   ИНН: {example.get('Идентификационный номер налогоплательщика', 'N/A')}")
                print(f"   Адрес: {example.get('Адрес, указанный в едином государственном реестре юридических лиц', 'N/A')[:80]}...")
            
            return mfo_list
            
        except Exception as e:
            print(f"❌ Ошибка парсинга справочника МФО: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def search_bank(self, banks: List[Dict], query: str) -> List[Dict]:
        """
        Поиск банка в справочнике по названию или ОГРН
        
        Args:
            banks: Список банков из справочника
            query: Строка поиска (название или ОГРН)
        
        Returns:
            Список найденных банков
        """
        query_lower = query.lower()
        results = []
        
        for bank in banks:
            # Ищем в названии
            name = bank.get('bnk_name', '')
            if name and query_lower in name.lower():
                results.append(bank)
                continue
            
            # Ищем по ОГРН
            ogrn = bank.get('ogrn', '')
            if ogrn and query in str(ogrn):
                results.append(bank)
        
        return results
    
    def get_bank_info_by_name(self, banks: List[Dict], bank_name: str) -> Optional[Dict]:
        """
        Получить информацию о банке по названию из справочника
        
        Args:
            banks: Список банков из XLSX
            bank_name: Название банка для поиска
        
        Returns:
            Словарь с информацией о банке или None
        """
        # Нормализуем название для поиска
        search_name = bank_name.lower().replace('пао', '').replace('ао', '').replace('оао', '').replace('ооо', '').replace('«', '').replace('»', '').strip()
        
        for bank in banks:
            bank_full_name = bank.get('bnk_name', '').lower()
            
            if search_name in bank_full_name or bank_full_name in search_name:
                return {
                    'название': bank.get('bnk_name', ''),
                    'адрес': bank.get('bnk_addr', ''),
                    'огрн': bank.get('ogrn', ''),
                    'рег_номер': bank.get('cregnum', ''),
                    'дата_регистрации': bank.get('reg_date', ''),
                    'статус_лицензии': bank.get('lic_status', '')
                }
        
        return None
    
    def normalize_address_with_gpt(self, long_address: str) -> str:
        """
        Нормализует длинный адрес из XLSX ЦБ к короткому формату через GPT-4o-mini
        
        Args:
            long_address: Длинный адрес (например: "Российская Федерация, город Москва, улица Вавилова, дом 19")
        
        Returns:
            Короткий адрес (например: "117312, г. Москва, ул. Вавилова, д. 19")
        """
        from openai import OpenAI
        import os
        
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        
        prompt = f"""Преобразуй адрес в короткий юридический формат для документов.

ДЛИННЫЙ АДРЕС: {long_address}

ПРАВИЛА:
1. Добавь индекс в начало (если нет - оставь без индекса)
2. Сократи: "город" → "г.", "улица" → "ул.", "дом" → "д.", "корпус" → "к.", "строение" → "стр.", "литера" → "лит."
3. Удали "Российская Федерация", "вн.тер.г.", "муниципальный округ"
4. Формат: "индекс, область/край (если есть), г. Город, ул. Улица, д. X, к. Y, стр. Z"
5. Убери лишние пробелы

ПРИМЕРЫ:
Вход: "Российская Федерация, город Москва, улица Вавилова, дом 19"
Выход: "117312, г. Москва, ул. Вавилова, д. 19"

Вход: "191144, г. Санкт-Петербург, Дегтярный переулок, д.11. лит. А"
Выход: "191144, г. Санкт-Петербург, Дегтярный пер., д. 11, лит. А"

Верни ТОЛЬКО короткий адрес, без объяснений."""

        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
            )
            normalized = response.choices[0].message.content.strip()
            return normalized
        except Exception as e:
            print(f"⚠️  Ошибка нормализации адреса: {e}")
            return long_address  # Возвращаем исходный адрес при ошибке
    
    def normalize_bank_name_from_xlsx(self, raw_name: str) -> str:
        """
        Нормализует название банка из XLSX к правильному формату
        
        Примеры:
            АКБ "Русский Трастовый Банк" (АО) → АО АКБ «Русский Трастовый Банк»
            ООО КБ "Альтайкапиталбанк" (ООО) → ООО КБ «Альтайкапиталбанк»
            
        Args:
            raw_name: Исходное название из XLSX
            
        Returns:
            Нормализованное название с елочками и правильным порядком ОПФ
        """
        import re
        
        # Убираем лишние пробелы
        name = raw_name.strip()
        
        # Извлекаем ОПФ из скобок в конце: (ПАО), (АО), (ООО) и т.д.
        opf_match = re.search(r'\(([А-ЯЁ]+)\)\s*$', name)
        opf_suffix = ""
        if opf_match:
            opf_suffix = opf_match.group(1)
            # Удаляем из названия
            name = re.sub(r'\s*\([А-ЯЁ]+\)\s*$', '', name)
        
        # Извлекаем ОПФ из начала: ПАО, АО, ООО и т.д.
        opf_prefix_match = re.match(r'^([А-ЯЁ]+)\s+', name)
        opf_prefix = ""
        if opf_prefix_match:
            potential_opf = opf_prefix_match.group(1)
            if potential_opf in ['ПАО', 'АО', 'ООО', 'ОАО', 'ЗАО', 'НКО', 'КБ', 'АКБ']:
                opf_prefix = potential_opf
                # Удаляем из названия
                name = re.sub(r'^[А-ЯЁ]+\s+', '', name, count=1)
        
        # Определяем финальный ОПФ (приоритет - из скобок)
        final_opf = opf_suffix if opf_suffix else opf_prefix
        
        # Заменяем прямые кавычки на елочки
        name = name.replace('"', '«').replace('"', '»')
        
        # Если кавычки уже есть - оставляем как есть
        # Если нет - добавляем вокруг основного названия
        if '«' not in name and '»' not in name:
            name = f'«{name}»'
        
        # Собираем итоговое название: ОПФ + название
        if final_opf:
            result = f'{final_opf} {name}'
        else:
            result = name
        
        return result

    def update_bank_registry_addresses(self, current_registry: Dict[str, Dict], banks: List[Dict]) -> Dict[str, Dict]:
        """
        Обновляет BANK_REGISTRY: берет ВСЕ банки из XLSX + добавляет ИНН из локальной базы
        
        Логика:
        1. Берем ВСЕ банки из XLSX (ключ = ОГРН, значение = название + адрес)
        2. Нормализуем названия банков (елочки, правильный порядок ОПФ)
        3. Ищем ИНН в локальной базе по названию банка
        4. Если ИНН найден и название совпадает - добавляем его
        
        Args:
            current_registry: Локальный BANK_REGISTRY {ИНН: {адрес, название}}
            banks: Список банков из XLSX справочника ЦБ РФ
        
        Returns:
            Полный обновленный словарь: {ОГРН: {название, адрес, инн (опционально)}}
        """
        updated_registry = {}
        added_count = 0
        inn_matched_count = 0
        
        print(f"\n[UPDATE] Обрабатываю {len(banks)} банков из справочника ЦБ РФ...")
        
        for bank in banks:
            # Получаем ОГРН как ключ
            ogrn = bank.get('ogrn', '').strip()
            if not ogrn:
                continue
            
            # Получаем название и адрес из XLSX (формат ЦБ РФ)
            raw_name = bank.get('bnk_name', '').strip()
            if not raw_name:
                continue
            
            # Нормализуем название (елочки, правильный порядок ОПФ)
            normalized_name = self.normalize_bank_name_from_xlsx(raw_name)
            
            # Получаем адрес из XLSX (используем как есть, без нормализации)
            raw_address = bank.get('bnk_addr', '').strip()
            
            # Ищем ИНН в локальной базе по названию
            matched_inn = None
            for inn, data in current_registry.items():
                local_name = data['название'].lower()
                xlsx_name = normalized_name.lower()
                
                # Упрощенное сравнение: убираем ОПФ и кавычки
                local_clean = local_name.replace('пао', '').replace('ао', '').replace('оао', '').replace('ооо', '').replace('«', '').replace('»', '').strip()
                xlsx_clean = xlsx_name.replace('пао', '').replace('ао', '').replace('оао', '').replace('ооо', '').replace('«', '').replace('»', '').strip()
                
                if local_clean in xlsx_clean or xlsx_clean in local_clean:
                    matched_inn = inn
                    inn_matched_count += 1
                    break
            
            # Добавляем банк с ОГРН как ключ
            updated_registry[ogrn] = {
                'название': normalized_name,
                'адрес': raw_address,
                'инн': matched_inn if matched_inn else ''  # ИНН из локальной базы (если нашли)
            }
            added_count += 1
        
        print(f"\n✅ Обновление завершено:")
        print(f"   Всего банков из XLSX: {len(updated_registry)}")
        print(f"   Из них с ИНН из локальной базы: {inn_matched_count}")
        print(f"   Без ИНН: {len(updated_registry) - inn_matched_count}")
        
        return updated_registry


def test_download_and_parse():
    """Тест скачивания и парсинга справочников"""
    
    print("=" * 80)
    print("📚 ТЕСТ: Скачивание и парсинг XLSX-справочников ЦБ РФ")
    print("=" * 80)
    
    registry = CBRExcelRegistry()
    
    # Скачиваем справочники
    banks_file = registry.download_banks_registry()
    mfo_file = registry.download_mfo_registry()
    
    banks_data = []
    mfo_data = []
    
    # Парсим банки
    if banks_file and banks_file.exists():
        banks_data = registry.parse_banks_registry(banks_file)
        
        if banks_data:
            print("\n📊 Пример данных банка:")
            print("-" * 80)
            example = banks_data[0]
            for key, value in list(example.items())[:10]:  # Показываем первые 10 полей
                print(f"   {key}: {value}")
    
    # Парсим МФО
    if mfo_file and mfo_file.exists():
        mfo_data = registry.parse_mfo_registry(mfo_file)
        
        if mfo_data:
            print("\n📊 Пример данных МФО:")
            print("-" * 80)
            example = mfo_data[0]
            for key, value in list(example.items())[:10]:
                print(f"   {key}: {value}")
    
    return banks_data, mfo_data


def test_search_in_registry():
    """Тест поиска банков в справочнике"""
    
    print("\n" + "=" * 80)
    print("🔍 ТЕСТ: Поиск банков в справочнике")
    print("=" * 80)
    
    registry = CBRExcelRegistry()
    
    # Проверяем, есть ли скачанный файл
    banks_file = registry.DATA_DIR / "banks_registry.xlsx"
    
    if not banks_file.exists():
        print("⚠️ Сначала нужно скачать справочник")
        banks_file = registry.download_banks_registry()
        if not banks_file:
            return
    
    # Парсим
    banks_data = registry.parse_banks_registry(banks_file)
    
    if not banks_data:
        print("❌ Нет данных для поиска")
        return
    
    # Тестовые запросы
    test_queries = ["Сбербанк", "Альфа", "ВТБ", "Тинькофф", "7707083893"]
    
    for query in test_queries:
        print(f"\n📌 Поиск: '{query}'")
        print("-" * 80)
        
        results = registry.search_bank(banks_data, query)
        
        if results:
            print(f"✅ Найдено: {len(results)} результат(ов)")
            for i, bank in enumerate(results[:3], 1):  # Показываем первые 3
                name = (bank.get('Сокращенное наименование') or 
                       bank.get('Наименование', 'N/A'))
                inn = bank.get('ИНН', 'N/A')
                address = (bank.get('Юридический адрес') or 
                          bank.get('Адрес', 'N/A'))
                
                print(f"\n   {i}. {name}")
                print(f"      ИНН: {inn}")
                print(f"      Адрес: {address[:100]}..." if len(address) > 100 else f"      Адрес: {address}")
        else:
            print("❌ Ничего не найдено")


def test_update_addresses():
    """Тест обновления адресов в существующем BANK_REGISTRY"""
    
    print("\n" + "=" * 80)
    print("🔄 ТЕСТ: Обновление адресов из актуального справочника ЦБ")
    print("=" * 80)
    
    registry = CBRExcelRegistry()
    banks_file = registry.DATA_DIR / "banks_registry.xlsx"
    
    if not banks_file.exists():
        print("⚠️ Сначала нужно скачать справочник")
        banks_file = registry.download_banks_registry()
        if not banks_file:
            return
    
    banks_data = registry.parse_banks_registry(banks_file)
    
    if not banks_data:
        print("❌ Нет данных")
        return
    
    # Примеры из текущего BANK_REGISTRY для теста
    sample_registry = {
        "7707083893": {"адрес": "Старый адрес", "название": "ПАО Сбербанк"},
        "7728168971": {"адрес": "Старый адрес", "название": "АО «АЛЬФА-БАНК»"},
        "7702070139": {"адрес": "Старый адрес", "название": "Банк ВТБ (ПАО)"},
    }
    
    print("\n📋 Тестовый реестр (3 банка):")
    for inn, data in sample_registry.items():
        print(f"   {data['название']} (ИНН: {inn})")
    
    # Обновляем адреса
    print("\n🔍 Поиск актуальных адресов в справочнике ЦБ...")
    updated_registry = registry.update_bank_registry_addresses(sample_registry, banks_data)
    
    print("\n📊 Результат обновления:")
    print("-" * 80)
    for inn, data in updated_registry.items():
        print(f"\n{data['название']} (ИНН: {inn})")
        print(f"Адрес: {data['адрес'][:100]}..." if len(data['адрес']) > 100 else f"Адрес: {data['адрес']}")


def test_compare_with_current():
    """Сравнение с текущим BANK_REGISTRY в processor.py"""
    
    print("\n" + "=" * 80)
    print("🔄 ТЕСТ: Сравнение со старым реестром")
    print("=" * 80)
    
    # TODO: Можно добавить загрузку текущего BANK_REGISTRY из processor.py
    # и сравнение с новыми данными
    
    print("ℹ️ Функция сравнения будет реализована при необходимости")
    print("Позволит увидеть:")
    print("  - Новые банки")
    print("  - Измененные адреса")
    print("  - Закрытые банки")


if __name__ == "__main__":
    print("\n" + "🚀 " + "=" * 76)
    print("📊 РАБОТА С XLSX-СПРАВОЧНИКАМИ ЦБ РФ")
    print("=" * 78 + "\n")
    
    print("📖 Источники данных:")
    print("   1. Банки: https://cbr.ru/banking_sector/credit/FullCoList/")
    print("   2. МФО: https://cbr.ru/vfs/finmarkets/files/supervision/list_MFO.xlsx")
    print()
    
    try:
        # Тест 1: Скачивание и парсинг
        banks_data, mfo_data = test_download_and_parse()
        
        # Тест 2: Поиск
        if banks_data:
            test_search_in_registry()
        
        # Тест 3: Обновление адресов
        if banks_data:
            test_update_addresses()
        
        # Тест 4: Сравнение
        test_compare_with_current()
        
        print("\n" + "=" * 80)
        print("✅ Все тесты завершены")
        print("=" * 80)
        print("\n💡 РЕКОМЕНДАЦИИ:")
        print("   1. Файлы сохранены в папку 'cbr_data/'")
        print("   2. Справочник содержит актуальные адреса всех действующих банков")
        print("   3. ИНН используются из локального BANK_REGISTRY в processor.py")
        print("   4. Адреса можно автоматически обновлять из XLSX-справочника")
        print("   5. Рекомендуется обновлять раз в месяц или при необходимости")
        print()
        
    except Exception as e:
        print(f"\n❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
